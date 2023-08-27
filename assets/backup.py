import serial
import re
import json
from datetime import datetime, time
import locale
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
#sudo chmod a+rw /dev/ttyACM0


port = '/dev/ttyACM0'  # Substitua 'COMX' pela porta serial do Arduino

# Configura o idioma para português do Brasil
locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')

# Função para buscar e exibir informações do aluno pelo ID
def exibir_informacoes_aluno(id, sheet):
    with open('alunos.json') as file:
        data = json.load(file)
        alunos = data['alunos']
        
        for aluno in alunos:
            if aluno['id'] == id:
                row = sheet.max_row + 1  # Obtém a próxima linha vazia
                sheet.cell(row=row, column=1, value=aluno["id"])
                sheet.cell(row=row, column=2, value=aluno["nome_completo"])
                sheet.cell(row=row, column=3, value=aluno["curso"])
                sheet.cell(row=row, column=4, value=aluno["RA"])
                print(f'Dados do aluno ID {id} registrados.')
                return
                
        sheet.append([f'Aluno com ID {id} não encontrado.'])

# Função para ler informações sobre a aula, horário e professor
def ler_informacoes_aula(sheet):
    with open('aula.json') as file:
        data = json.load(file)
        aulas = data['aulas']
        
        horario_atual = datetime.now().time()
        dia_semana_atual = datetime.now().strftime("%A")

        for aula in aulas:
            horario_inicio = time.fromisoformat(aula['horario_inicio'])
            horario_fim = time.fromisoformat(aula['horario_fim'])
            
            if verificar_dia_semana_aula(dia_semana_atual, aula['dia_semana']) and verificar_horario_aula(horario_atual, horario_inicio, horario_fim):
                row = sheet.max_row + 1  # Obtém a próxima linha vazia
                sheet.cell(row=row, column=5, value=aula["nome_aula"])
                sheet.cell(row=row, column=6, value=aula["dia_semana"])
                sheet.cell(row=row, column=7, value=datetime.now().strftime("%H:%M:%S"))
                sheet.cell(row=row, column=8, value=f'{aula["horario_inicio"]} - {aula["horario_fim"]}')
                sheet.cell(row=row, column=9, value=aula["nome_professor"])
                print('Dados da aula registrados.')
                return
            
        sheet.append(['Nenhuma aula encontrada no momento.'])

# Função para verificar se o dia da semana da aula corresponde ao dia da semana atual
def verificar_dia_semana_aula(dia_semana_atual, dia_semana_aula):
    return dia_semana_atual.lower() == dia_semana_aula.lower()

# Função para verificar se o horário atual está dentro do intervalo de horário da aula
def verificar_horario_aula(horario_atual, horario_inicio, horario_fim):
    return horario_inicio <= horario_atual <= horario_fim

# Abre a porta serial
with serial.Serial(port, 9600) as ser:
    wb = Workbook()
    sheet = wb.active

    # Define os cabeçalhos da tabela
    headers = [
        "ID do Aluno",
        "Nome Completo",
        "Curso",
        "RA",
        "Nome da Aula",
        "Dia da Semana",
        "Horário de Registro",
        "Horário da Aula",
        "Nome do Professor"
    ]
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        cell = sheet[f"{column_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    while True:
        line = ser.readline().decode(errors='ignore').strip()

        if line == "No finger detected":
            continue  # Pula para a próxima iteração do loop

        # Procura por um padrão de ID na linha
        match = re.search(r'Found ID #(\d+) with confidence', line)
        if match:
            id = int(match.group(1))  # Obtém o ID correspondente como inteiro

            # Chama a função para exibir as informações do aluno
            exibir_informacoes_aluno(id, sheet)

            # Verifica o horário da aula
            ler_informacoes_aula(sheet)

            # Salva a planilha em um arquivo
            wb.save('dados_alunos.xlsx')